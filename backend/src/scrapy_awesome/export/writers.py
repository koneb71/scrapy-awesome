"""Write rows to JSON / JSONL / CSV / XLSX.

Column order: recipe field order if known, else first-seen order; implicit `_` columns go last
(and `_provenance` is dropped unless `include_meta=True`). Nested values are JSON-encoded for
CSV/XLSX; lists join with `; ` when every element is a scalar.
"""

from __future__ import annotations

import csv
import json
from collections.abc import Iterable
from pathlib import Path
from typing import Any

FORMATS = ("json", "jsonl", "csv", "xlsx")
_META_ORDER = ("_url", "_page_url", "_fetched_at", "_tier", "_provenance")


def _cell(v: Any) -> Any:
    if v is None or isinstance(v, str | int | float | bool):
        return v
    if isinstance(v, list) and all(isinstance(x, str | int | float | bool) or x is None for x in v):
        return "; ".join("" if x is None else str(x) for x in v)
    return json.dumps(v, ensure_ascii=False, default=str)


def column_order(
    rows: Iterable[dict[str, Any]], *, fields: list[str] | None = None, include_meta: bool = True
) -> list[str]:
    seen: list[str] = list(fields or [])
    meta: list[str] = []
    for r in rows:
        for k in r:
            if k.startswith("_"):
                if k not in meta:
                    meta.append(k)
            elif k not in seen:
                seen.append(k)
    if include_meta:
        ordered_meta = [m for m in _META_ORDER if m in meta] + [
            m for m in meta if m not in _META_ORDER
        ]
        seen += [m for m in ordered_meta if m != "_provenance"]
    return seen


def read_jsonl(path: Path) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    with Path(path).open(encoding="utf-8") as fh:
        for line in fh:
            line = line.strip()
            if line:
                rows.append(json.loads(line))
    return rows


def export_rows(
    rows: list[dict[str, Any]],
    out: Path,
    *,
    fmt: str,
    fields: list[str] | None = None,
    include_meta: bool = True,
    sheet_name: str = "items",
) -> Path:
    out = Path(out)
    out.parent.mkdir(parents=True, exist_ok=True)
    fmt = fmt.lower()
    if fmt not in FORMATS:
        raise ValueError(f"unknown format {fmt!r}; expected one of {FORMATS}")
    cols = column_order(rows, fields=fields, include_meta=include_meta)

    if fmt == "jsonl":
        with out.open("w", encoding="utf-8") as fh:
            for r in rows:
                fh.write(
                    json.dumps(
                        {k: r.get(k) for k in cols if k in r}, ensure_ascii=False, default=str
                    )
                    + "\n"
                )
        return out
    if fmt == "json":
        data = [{k: r.get(k) for k in cols if k in r} for r in rows]
        out.write_text(
            json.dumps(data, ensure_ascii=False, indent=2, default=str), encoding="utf-8"
        )
        return out
    if fmt == "csv":
        with out.open(
            "w", encoding="utf-8-sig", newline=""
        ) as fh:  # BOM → Excel opens UTF-8 correctly
            w = csv.DictWriter(fh, fieldnames=cols, extrasaction="ignore")
            w.writeheader()
            for r in rows:
                w.writerow({k: _cell(r.get(k)) for k in cols})
        return out
    # xlsx
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31] or "items"
    ws.append(cols)
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="EEEEEE")
    for r in rows:
        ws.append([_cell(r.get(k)) for k in cols])
    for i, col in enumerate(cols, start=1):
        width = max([len(str(col))] + [len(str(_cell(r.get(col)) or "")) for r in rows[:200]])
        ws.column_dimensions[get_column_letter(i)].width = min(max(10, width + 2), 60)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(out)
    return out


def export_jsonl_file(
    items_path: Path,
    *,
    fmt: str,
    out: Path | None = None,
    fields: list[str] | None = None,
    include_meta: bool = True,
) -> Path:
    items_path = Path(items_path)
    rows = read_jsonl(items_path)
    dest = out or items_path.with_name(f"items.{fmt}")
    return export_rows(rows, dest, fmt=fmt, fields=fields, include_meta=include_meta)
